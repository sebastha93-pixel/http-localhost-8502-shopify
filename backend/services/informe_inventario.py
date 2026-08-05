"""informe_inventario.py — el inventario de telas e insumos, para imprimir.

POR QUÉ EXISTE (2026-08-05, pedido de Sebastián): los datos estaban en pantalla
pero no había forma de sacarlos. Un inventario sirve para caminar la bodega y
contar, y eso se hace con una hoja en la mano.

DOS DECISIONES DE DISEÑO QUE IMPORTAN:

1. COLUMNA "CONTEO FÍSICO" EN BLANCO. El informe no es para leerlo, es para
   cotejarlo: al lado de lo que el sistema cree, un espacio para escribir lo que
   de verdad hay. Sin esa columna el papel solo sirve para mirar; con ella cierra
   el ciclo y de ahí salen los ajustes.

2. EL COSTO SE OMITE SI NO SE TIENE EL PERMISO. `inventario_resumen` ya esconde
   `valor_estimado` a quien no tiene `produccion_costos`, y el PDF respeta lo
   mismo: un informe que se imprime y se deja sobre una mesa no puede filtrar
   costos que en pantalla están tapados.

Formatos: PDF (para imprimir) y XLSX (para trabajar los números). Los dos salen de
los mismos datos, así que no pueden discrepar.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta, timezone
from typing import Optional

log = logging.getLogger(__name__)

_TZ_BOGOTA = timezone(timedelta(hours=-5))


def _ahora() -> str:
    return datetime.now(_TZ_BOGOTA).strftime("%Y-%m-%d %H:%M")


def _datos(*, con_costos: bool) -> dict:
    """Telas agrupadas + insumos, ya ordenados como se van a imprimir."""
    from backend.services import produccion as pr

    telas = list(pr.inventario_resumen() or [])
    telas.sort(key=lambda t: ((t.get("descripcion_tela") or "").upper(),
                              (t.get("tono") or "")))
    insumos = [i for i in (pr.listar_insumos() or []) if i.get("activo", True)]
    insumos.sort(key=lambda i: ((i.get("categoria") or "").upper(),
                                (i.get("nombre") or "").upper()))
    if not con_costos:
        telas = [{k: v for k, v in t.items() if k != "valor_estimado"} for t in telas]
        insumos = [{k: v for k, v in i.items() if k != "costo_unitario"} for i in insumos]
    return {"telas": telas, "insumos": insumos, "generado": _ahora(),
            "con_costos": con_costos}


def _stock_insumo(i: dict) -> float:
    """El insumo trae varios nombres de stock según su antigüedad."""
    for k in ("stock_disponible", "cantidad_disponible", "stock_inicial"):
        v = i.get(k)
        if v is not None:
            try:
                return float(v)
            except Exception:
                continue
    return 0.0


# ── PDF ──────────────────────────────────────────────────────────────────────

def pdf(*, tipo: str = "ambos", con_costos: bool = False) -> bytes:
    """Informe imprimible. `tipo`: telas | insumos | ambos."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)

    d = _datos(con_costos=con_costos)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm,
                            title="Inventario MALE'DENIM")
    est = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=est["Heading1"], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=est["Normal"], fontSize=8,
                         textColor=colors.HexColor("#666666"), spaceAfter=8)
    # Las celdas de texto largo van como Paragraph, no como str: un str se
    # DESBORDA sobre la columna vecina sin avisar (la composición de las telas
    # tiene hasta 45 caracteres y se montaba encima de "Rollos"). El Paragraph
    # respeta el ancho y parte la línea.
    celda = ParagraphStyle("celda", parent=est["Normal"], fontSize=7,
                           leading=8.2, spaceBefore=0, spaceAfter=0)
    partes = []

    # Rejilla común. La última columna va vacía a propósito — es para escribir.
    def _estilo(n_filas: int, col_num: list[int]) -> TableStyle:
        cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3550")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Franjas: contar 90 renglones en una hoja blanca es como se salta uno.
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f4f6f8")]),
        ]
        for c in col_num:
            cmds.append(("ALIGN", (c, 1), (c, -1), "RIGHT"))
        return TableStyle(cmds)

    if tipo in ("telas", "ambos"):
        t = d["telas"]
        tot_m = sum(float(x.get("metros_disponible") or 0) for x in t)
        tot_r = sum(int(x.get("num_rollos") or 0) for x in t)
        partes.append(Paragraph("Inventario de telas", h))
        partes.append(Paragraph(
            f"MALE'DENIM · generado {d['generado']} · {len(t)} referencias · "
            f"{tot_r} rollos · {tot_m:,.1f} m disponibles", sub))
        cab = ["Tela", "Tono", "Composición", "Rollos", "Disponible (m)"]
        if con_costos:
            cab.append("Valor est.")
        cab.append("Conteo físico")
        filas = [cab]
        for x in t:
            f = [
                Paragraph(str(x.get("descripcion_tela") or "—")[:30], celda),
                str(x.get("tono") or "—")[:10],
                Paragraph(str(x.get("composicion") or "—")[:60], celda),
                str(int(x.get("num_rollos") or 0)),
                f"{float(x.get('metros_disponible') or 0):,.1f}",
            ]
            if con_costos:
                f.append(f"${float(x.get('valor_estimado') or 0):,.0f}")
            f.append("")          # el espacio para escribir a mano
            filas.append(f)
        total = ["TOTAL", "", "", str(tot_r), f"{tot_m:,.1f}"]
        if con_costos:
            total.append(f"${sum(float(x.get('valor_estimado') or 0) for x in t):,.0f}")
        total.append("")
        filas.append(total)
        anchos = ([34, 14, 46, 13, 22] + ([22] if con_costos else []) + [24])
        tb = Table(filas, colWidths=[a * mm for a in anchos], repeatRows=1)
        st = _estilo(len(filas), [3, 4] + ([5] if con_costos else []))
        st.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
        st.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8ecf0"))
        tb.setStyle(st)
        partes.append(tb)

    if tipo == "ambos":
        partes.append(PageBreak())

    if tipo in ("insumos", "ambos"):
        ins = d["insumos"]
        partes.append(Paragraph("Inventario de insumos", h))
        partes.append(Paragraph(
            f"MALE'DENIM · generado {d['generado']} · {len(ins)} insumos activos", sub))
        cab = ["Código", "Insumo", "Categoría", "Unidad", "Stock"]
        if con_costos:
            cab.append("Costo unit.")
        cab.append("Conteo físico")
        filas = [cab]
        cat_actual = None
        for x in ins:
            cat = (x.get("categoria") or "—")
            # Subtítulo por categoría: separa confección de terminación, que es
            # como está repartida la bodega.
            if cat != cat_actual:
                cat_actual = cat
                fila = [cat.title(), "", "", "", ""] + ([""] if con_costos else []) + [""]
                filas.append(fila)
            f = [
                str(x.get("codigo") or "—")[:16],
                Paragraph(str(x.get("nombre") or "—")[:44], celda),
                "", "",
                f"{_stock_insumo(x):,.1f}",
            ]
            f[2] = ""            # la categoría ya va en el subtítulo
            f[3] = str(x.get("unidad") or "—")[:8]
            if con_costos:
                f.append(f"${float(x.get('costo_unitario') or 0):,.0f}")
            f.append("")
            filas.append(f)
        anchos = ([26, 50, 4, 14, 22] + ([22] if con_costos else []) + [24])
        tb = Table(filas, colWidths=[a * mm for a in anchos], repeatRows=1)
        st = _estilo(len(filas), [4] + ([5] if con_costos else []))
        # Resaltar los renglones de categoría
        for idx, fila in enumerate(filas):
            if idx and fila[1] == "" and fila[4] == "":
                st.add("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#dde3ea"))
                st.add("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold")
                st.add("SPAN", (0, idx), (3, idx))
        tb.setStyle(st)
        partes.append(tb)

    partes.append(Spacer(1, 6 * mm))
    partes.append(Paragraph(
        "Conteo hecho por: ______________________________   "
        "Fecha: ____________   Firma: ______________________",
        ParagraphStyle("pie", parent=est["Normal"], fontSize=8)))
    doc.build(partes)
    return buf.getvalue()


# ── XLSX ─────────────────────────────────────────────────────────────────────

def xlsx(*, con_costos: bool = False) -> bytes:
    """Mismo inventario, en Excel: una hoja de telas y una de insumos."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    d = _datos(con_costos=con_costos)
    wb = Workbook()
    azul = PatternFill("solid", start_color="1F3550")
    blanco = Font(bold=True, color="FFFFFF")

    def _cabecera(ws, cab):
        ws.append(cab)
        for c in ws[1]:
            c.fill = azul; c.font = blanco
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Telas"
    cab = ["Tela", "Tono", "Composición", "Rollos", "Disponible (m)", "Ingresado (m)"]
    if con_costos:
        cab.append("Valor estimado")
    cab.append("Conteo físico")
    _cabecera(ws, cab)
    for x in d["telas"]:
        fila = [x.get("descripcion_tela"), x.get("tono"), x.get("composicion"),
                int(x.get("num_rollos") or 0),
                round(float(x.get("metros_disponible") or 0), 2),
                round(float(x.get("metros_inicial") or 0), 2)]
        if con_costos:
            fila.append(round(float(x.get("valor_estimado") or 0), 2))
        fila.append(None)        # columna para el conteo
        ws.append(fila)
    for col, w in zip("ABCDEFGH", (26, 10, 34, 8, 15, 14, 16, 14)):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Insumos")
    cab2 = ["Código", "Insumo", "Categoría", "Unidad", "Stock", "Stock mínimo"]
    if con_costos:
        cab2.append("Costo unitario")
    cab2.append("Conteo físico")
    _cabecera(ws2, cab2)
    for x in d["insumos"]:
        fila = [x.get("codigo"), x.get("nombre"), x.get("categoria"),
                x.get("unidad"), round(_stock_insumo(x), 2),
                float(x.get("stock_minimo") or 0)]
        if con_costos:
            fila.append(round(float(x.get("costo_unitario") or 0), 2))
        fila.append(None)
        ws2.append(fila)
    for col, w in zip("ABCDEFGH", (16, 34, 22, 10, 12, 13, 15, 14)):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
