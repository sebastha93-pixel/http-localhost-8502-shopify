"""
backend.services.informe_consultor — Informe director comercial automatizado.

Toma N conversaciones (típicamente últimos 7 días) y genera el informe A-I
con Claude Haiku 4.5: razones de pérdida, patrones, errores, 20 frases que
matan, 20 mejores prácticas, comparativa por asesora.

Uso típico:
    informe = generar_informe(days_back=7)
"""
from __future__ import annotations

import json
import os
from datetime import datetime, timedelta, timezone
from typing import Any
from zoneinfo import ZoneInfo

import anthropic

from backend.services import revenue_db as db


MODELO = "claude-haiku-4-5-20251001"
MAX_TOKENS = 16000
TZ_BOG = ZoneInfo("America/Bogota")
MAX_CONVERSACIONES = 80


SYSTEM_PROMPT = """Eres un director comercial senior con 20+ años de experiencia en \
ecommerce, ventas por WhatsApp, moda femenina y optimización de conversión.

Vas a analizar conversaciones reales entre asesoras de MALE Denim (marca colombiana \
de jeans premium para mujer) y sus clientas. Tu tarea NO es resumir. Es ENCONTRAR \
LAS FUGAS DE DINERO y producir un informe brutal que el director comercial pueda \
usar para tomar decisiones.

Contexto MALE Denim:
- Marca colombiana de jeans premium para mujer
- Productos: jeans, precios $149.900 – $329.800 COP
- Canales: WhatsApp Business + Instagram
- Logística: Melonn

Tono: director comercial directo, sin endulzar. Cuantifica impacto $. Cita textualmente.

Output SIEMPRE en JSON válido siguiendo el esquema indicado por el usuario."""


def _resumir_conversacion(conv: dict) -> str:
    msgs = conv.get("messages", []) or []
    msgs = msgs[-40:]
    lineas = []
    for m in msgs:
        sender = m.get("sender_type") or ""
        direccion = "ASESORA" if sender == "advisor" else "CLIENTA" if sender == "customer" else "SISTEMA"
        texto = (m.get("message_text") or "").strip().replace("\n", " ")
        if not texto:
            continue
        lineas.append(f"{direccion}: {texto[:300]}")
    return "\n".join(lineas)


def _fetch_conversaciones(
    days_back: int = 7,
    advisor_id: str | None = None,
    limit: int = MAX_CONVERSACIONES,
) -> list[dict]:
    sb = db._sb()
    if sb is None:
        return []
    desde = (datetime.now(timezone.utc) - timedelta(days=days_back)).isoformat()

    q = (
        sb.table("conversations")
        .select("conversation_id, lead_id, advisor_id, channel, status, last_message_at")
        .gte("last_message_at", desde)
        .order("last_message_at", desc=True)
        .limit(limit)
    )
    if advisor_id:
        q = q.eq("advisor_id", advisor_id)

    convs = q.execute().data or []
    if not convs:
        return []

    lead_ids = list({c["lead_id"] for c in convs if c.get("lead_id")})
    advisor_ids = list({c["advisor_id"] for c in convs if c.get("advisor_id")})
    leads_map: dict = {}
    advisors_map: dict = {}
    if lead_ids:
        for l in (sb.table("kommo_leads").select("lead_id,customer_name,status,lead_value").in_("lead_id", lead_ids).execute().data or []):
            leads_map[l["lead_id"]] = l
    if advisor_ids:
        for a in (sb.table("advisors").select("advisor_id,name").in_("advisor_id", advisor_ids).execute().data or []):
            advisors_map[a["advisor_id"]] = a

    for c in convs:
        lead = leads_map.get(c.get("lead_id"), {})
        adv = advisors_map.get(c.get("advisor_id"), {})
        c["customer_name"] = lead.get("customer_name") or "sin nombre"
        c["lead_status"] = lead.get("status") or "open"
        c["lead_value"] = lead.get("lead_value") or 0
        c["asesora_nombre"] = adv.get("name") or "Sin asignar"

        msgs = (
            sb.table("messages")
            .select("sender_type, message_text, sent_at")
            .eq("conversation_id", c["conversation_id"])
            .order("sent_at")
            .limit(60)
            .execute()
            .data
            or []
        )
        c["messages"] = msgs

    return convs


def _build_user_prompt(convs: list[dict]) -> str:
    bloques = []
    for i, c in enumerate(convs, 1):
        bloques.append(
            f"### CONV {i} — Lead {c.get('lead_id')} | Asesora: {c['asesora_nombre']} | "
            f"Cliente: {c['customer_name']} | "
            f"Status: {c['lead_status']} | Valor: ${c['lead_value']:,}\n"
            f"{_resumir_conversacion(c)}"
        )

    conversaciones_txt = "\n\n".join(bloques)

    return f"""Analiza las siguientes {len(convs)} conversaciones reales de MALE Denim \
y produce un informe consultor en JSON con esta estructura EXACTA:

{{
  "resumen": {{
    "total_conversaciones": int,
    "ganadas": int,
    "perdidas": int,
    "inconclusas": int,
    "conv_rate_pct": float,
    "valor_recuperable_estimado_cop": int,
    "diagnostico_general": "string (1 párrafo brutal)"
  }},
  "seccion_a_top10_razones_perdida": [
    {{"razon": "string", "impacto_cop": int, "frecuencia": int, "ejemplo_cita": "string"}}
  ],
  "seccion_b_patrones_perdidas": [
    {{"patron": "string", "cita_textual": "string", "frecuencia": int}}
  ],
  "seccion_c_errores_asesoras": [
    {{"error": "string", "cita_textual": "string", "asesora_ejemplo": "string"}}
  ],
  "seccion_d_preguntas_frecuentes_clientas": [
    {{"pregunta": "string", "frecuencia": int}}
  ],
  "seccion_e_diff_compra_vs_no": {{
    "ganadoras_hacen": ["string", ...],
    "perdidas_hacen": ["string", ...]
  }},
  "seccion_f_oportunidades": [
    {{"oportunidad": "string", "impacto_cop_mensual": int, "esfuerzo": "alto|medio|bajo"}}
  ],
  "seccion_g_plan_accion": [
    {{"accion": "string", "impacto_cop": int, "esfuerzo": "alto|medio|bajo", "responsable": "string", "plazo_dias": int}}
  ],
  "seccion_h_top20_frases_negativas": [
    {{"frase": "string", "por_que_mata_venta": "string"}}
  ],
  "seccion_i_top20_mejores_practicas": [
    {{"frase": "string", "por_que_funciona": "string"}}
  ],
  "comparativa_asesoras": [
    {{
      "asesora": "string",
      "conv_totales": int,
      "ganadas": int,
      "perdidas": int,
      "conv_rate_pct": float,
      "fortalezas": ["string", "string", "string"],
      "debilidades": ["string", "string", "string"],
      "recomendacion": "string"
    }}
  ]
}}

Reglas:
- Cuantifica impacto $ siempre que puedas
- Las citas textuales deben venir DE las conversaciones reales mostradas
- En H y I incluye exactamente 20 frases (si hay menos data, infiere y márcalas con "(inferido)")
- En A son exactamente 10 razones ordenadas por impacto $ (no por frecuencia)

CONVERSACIONES:

{conversaciones_txt}

Responde SOLO con el JSON, sin markdown, sin explicación adicional."""


def generar_informe(
    days_back: int = 7,
    advisor_id: str | None = None,
) -> dict[str, Any]:
    convs = _fetch_conversaciones(days_back=days_back, advisor_id=advisor_id)

    if not convs:
        return {
            "error": "Sin conversaciones en el periodo solicitado",
            "days_back": days_back,
            "advisor_id": advisor_id,
        }

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return {"error": "ANTHROPIC_API_KEY no configurada"}

    client = anthropic.Anthropic(api_key=api_key)
    prompt = _build_user_prompt(convs)

    resp = client.messages.create(
        model=MODELO,
        max_tokens=MAX_TOKENS,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = resp.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        if raw.startswith("json"):
            raw = raw[4:].strip()
        raw = raw.rsplit("```", 1)[0].strip()

    try:
        informe = json.loads(raw)
    except json.JSONDecodeError as e:
        return {
            "error": f"JSON inválido del modelo: {e}",
            "raw_preview": raw[:500],
        }

    informe["_metadata"] = {
        "generado_at": datetime.now(TZ_BOG).isoformat(),
        "conversaciones_analizadas": len(convs),
        "days_back": days_back,
        "advisor_id": advisor_id,
        "modelo": MODELO,
        "tokens_input": resp.usage.input_tokens,
        "tokens_output": resp.usage.output_tokens,
    }

    return informe


def guardar_informe(informe: dict[str, Any]) -> str | None:
    sb = db._sb()
    if sb is None:
        return None
    try:
        row = {
            "generado_at": informe["_metadata"]["generado_at"],
            "days_back": informe["_metadata"]["days_back"],
            "advisor_id": informe["_metadata"].get("advisor_id"),
            "conversaciones_analizadas": informe["_metadata"]["conversaciones_analizadas"],
            "tokens_input": informe["_metadata"]["tokens_input"],
            "tokens_output": informe["_metadata"]["tokens_output"],
            "informe_json": informe,
        }
        res = sb.table("revenue_informes").insert(row).execute()
        return res.data[0]["id"] if res.data else None
    except Exception as e:
        print(f"⚠️ guardar_informe falló: {e}")
        return None


# ─── Export XLSX ──────────────────────────────────────────────────────────────

def informe_a_xlsx(informe: dict[str, Any]) -> bytes:
    """Convierte el informe A-I a un .xlsx con una hoja por sección.
    Formato pensado para que el director comercial lo lea sin tocar nada:
    encabezados con fondo, columnas anchas, montos COP con separador de miles."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", start_color="1F2A44")  # navy Selvedge
    body_font = Font(size=10, name="Arial")
    wrap = Alignment(vertical="top", wrap_text=True)
    COP = '#,##0'

    def hoja(nombre: str, headers: list[str], filas: list[list[Any]],
             anchos: list[int], money_cols: tuple[int, ...] = ()) -> None:
        ws = wb.create_sheet(title=nombre[:31])
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
        for fila in filas:
            ws.append(fila)
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = wrap
                if cell.column in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = COP
        ws.freeze_panes = "A2"

    r = informe.get("resumen") or {}
    meta = informe.get("_metadata") or {}
    hoja("Resumen", ["Métrica", "Valor"], [
        ["Generado", (meta.get("generado_at") or "")[:16].replace("T", " ")],
        ["Días analizados", meta.get("days_back")],
        ["Conversaciones analizadas", meta.get("conversaciones_analizadas")],
        ["Total conversaciones", r.get("total_conversaciones")],
        ["Ganadas", r.get("ganadas")],
        ["Perdidas", r.get("perdidas")],
        ["Inconclusas", r.get("inconclusas")],
        ["Tasa de conversión %", r.get("conv_rate_pct")],
        ["Valor recuperable estimado (COP)", r.get("valor_recuperable_estimado_cop")],
        ["Diagnóstico", r.get("diagnostico_general")],
    ], [34, 90], money_cols=(2,))

    hoja("A Razones de pérdida", ["Razón", "Impacto COP", "Frecuencia", "Cita de ejemplo"],
         [[x.get("razon"), x.get("impacto_cop"), x.get("frecuencia"), x.get("ejemplo_cita")]
          for x in informe.get("seccion_a_top10_razones_perdida") or []],
         [45, 14, 11, 70], money_cols=(2,))

    hoja("B Patrones", ["Patrón", "Cita textual", "Frecuencia"],
         [[x.get("patron"), x.get("cita_textual"), x.get("frecuencia")]
          for x in informe.get("seccion_b_patrones_perdidas") or []],
         [50, 70, 11])

    hoja("C Errores asesoras", ["Error", "Cita textual", "Asesora"],
         [[x.get("error"), x.get("cita_textual"), x.get("asesora_ejemplo")]
          for x in informe.get("seccion_c_errores_asesoras") or []],
         [50, 70, 20])

    hoja("D Preguntas frecuentes", ["Pregunta", "Frecuencia"],
         [[x.get("pregunta"), x.get("frecuencia")]
          for x in informe.get("seccion_d_preguntas_frecuentes_clientas") or []],
         [80, 11])

    e = informe.get("seccion_e_diff_compra_vs_no") or {}
    gan, per = e.get("ganadoras_hacen") or [], e.get("perdidas_hacen") or []
    hoja("E Compra vs No compra", ["Las que COMPRAN — la asesora hace", "Las que NO compran — la asesora hace"],
         [[gan[i] if i < len(gan) else "", per[i] if i < len(per) else ""]
          for i in range(max(len(gan), len(per)) or 1)],
         [70, 70])

    hoja("F Oportunidades", ["Oportunidad", "Impacto COP mensual", "Esfuerzo"],
         [[x.get("oportunidad"), x.get("impacto_cop_mensual"), x.get("esfuerzo")]
          for x in informe.get("seccion_f_oportunidades") or []],
         [70, 18, 10], money_cols=(2,))

    hoja("G Plan de acción", ["Acción", "Impacto COP", "Esfuerzo", "Responsable", "Plazo (días)"],
         [[x.get("accion"), x.get("impacto_cop"), x.get("esfuerzo"), x.get("responsable"), x.get("plazo_dias")]
          for x in informe.get("seccion_g_plan_accion") or []],
         [65, 14, 10, 20, 11], money_cols=(2,))

    hoja("H Frases que matan", ["Frase", "Por qué mata la venta"],
         [[x.get("frase"), x.get("por_que_mata_venta")]
          for x in informe.get("seccion_h_top20_frases_negativas") or []],
         [70, 70])

    hoja("I Mejores prácticas", ["Frase", "Por qué funciona"],
         [[x.get("frase"), x.get("por_que_funciona")]
          for x in informe.get("seccion_i_top20_mejores_practicas") or []],
         [70, 70])

    hoja("Asesoras", ["Asesora", "Conv. totales", "Ganadas", "Perdidas", "Conv. %",
                      "Fortalezas", "Debilidades", "Recomendación"],
         [[x.get("asesora"), x.get("conv_totales"), x.get("ganadas"), x.get("perdidas"),
           x.get("conv_rate_pct"), " · ".join(x.get("fortalezas") or []),
           " · ".join(x.get("debilidades") or []), x.get("recomendacion")]
          for x in informe.get("comparativa_asesoras") or []],
         [20, 12, 10, 10, 10, 55, 55, 60])

    wb.remove(wb["Sheet"])  # hoja default vacía
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
